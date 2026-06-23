from __future__ import annotations

import shutil
import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.append(str(PROJECT_ROOT))

from src.utils.helpers import parse_source_urls

WORKBOOK_PATH = PROJECT_ROOT / "data" / "family_offices_final.xlsx"
BACKUP_PATH = PROJECT_ROOT / "data" / "family_offices_final.pre_production_backup.xlsx"

UNKNOWN_VALUES = {"", "unknown", "nan", "none", "null"}


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in UNKNOWN_VALUES else text


def sentence(label: str, value: Any) -> str:
    text = clean(value)
    return f"{label}: {text}." if text else ""


def join_location(row: dict[str, Any]) -> str:
    parts = [
        clean(row.get("headquarters_city")),
        clean(row.get("headquarters_state_region")),
        clean(row.get("country")),
    ]
    return ", ".join(part for part in parts if part)


def rows_by_header(sheet) -> list[dict[str, Any]]:
    headers = [cell.value for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, values, strict=False)))
    return rows


def sheet_headers(sheet) -> dict[str, int]:
    return {str(cell.value): index for index, cell in enumerate(sheet[1], start=1)}


def source_urls(source_rows: list[dict[str, Any]]) -> list[str]:
    urls: list[str] = []
    for row in source_rows:
        for url in parse_source_urls(clean(row.get("source_url"))):
            if url not in urls:
                urls.append(url)
    return urls


def normalize_source_log_urls(sheet) -> int:
    headers = sheet_headers(sheet)
    if "source_url" not in headers:
        return 0
    updated = 0
    source_url_col = headers["source_url"]
    for row_number in range(2, sheet.max_row + 1):
        raw_value = sheet.cell(row=row_number, column=source_url_col).value
        urls = parse_source_urls(raw_value)
        if len(urls) == 1 and clean(raw_value) != urls[0]:
            sheet.cell(row=row_number, column=source_url_col).value = urls[0]
            updated += 1
        elif len(urls) > 1:
            sheet.cell(row=row_number, column=source_url_col).value = urls[0]
            updated += 1
    return updated


def build_entity_text(row: dict[str, Any]) -> str:
    name = clean(row.get("family_office_name"))
    family_type = clean(row.get("family_office_type"))
    location = join_location(row)
    principal = clean(row.get("principal_or_family"))

    pieces = []
    if name and family_type:
        pieces.append(f"{name} is listed as a {family_type}.")
    if location:
        pieces.append(f"Location: {location}.")
    if principal:
        pieces.append(f"Associated principal or family: {principal}.")
    pieces.extend(
        part
        for part in [
            sentence("Investment focus", row.get("investment_focus")),
            sentence("Asset classes", row.get("asset_classes")),
            sentence("Known portfolio or investments", row.get("known_portfolio_or_investments")),
            sentence("Recent activity or signal", row.get("recent_activity_or_signal")),
        ]
        if part
    )

    aum = clean(row.get("aum_estimate"))
    aum_source = clean(row.get("aum_source"))
    if aum:
        pieces.append(f"AUM or asset metric: {aum}." + (f" Source: {aum_source}." if aum_source else ""))
    return " ".join(pieces)


def build_investment_text(row: dict[str, Any]) -> str:
    name = clean(row.get("family_office_name"))
    pieces = [f"{name} investment profile."] if name else []
    pieces.extend(
        part
        for part in [
            sentence("Investment focus", row.get("investment_focus")),
            sentence("Asset classes", row.get("asset_classes")),
            sentence("Normalized asset classes", row.get("normalized_asset_classes")),
            sentence("Investing sectors", row.get("investing_sectors")),
            sentence("Stage preference", row.get("stage_preference")),
            sentence("Geography preference", row.get("geography_preference")),
            sentence("Known portfolio or investments", row.get("known_portfolio_or_investments")),
        ]
        if part
    )
    return " ".join(pieces)


def build_source_text(row: dict[str, Any], sources: list[dict[str, Any]]) -> str:
    name = clean(row.get("family_office_name"))
    pieces = [f"{name} source coverage."] if name else []

    source_types = sorted({clean(source.get("source_type")) for source in sources if clean(source.get("source_type"))})
    claims = []
    fields = []
    for source in sources:
        claim = clean(source.get("claim_validated"))
        if claim and claim not in claims:
            claims.append(claim)
        supported = clean(source.get("fields_supported"))
        if supported and supported not in fields:
            fields.append(supported)

    if source_types:
        pieces.append(f"Source types: {'; '.join(source_types)}.")
    if claims:
        pieces.append(f"Public sources support: {'; '.join(claims)}.")
    if fields:
        pieces.append(f"Covered fields: {'; '.join(fields)}.")
    return " ".join(pieces)


def build_activity_text(row: dict[str, Any], activity: dict[str, Any] | None) -> str:
    name = clean(row.get("family_office_name"))
    pieces = [f"{name} recent activity."] if name else []
    if activity:
        pieces.extend(
            part
            for part in [
                sentence("Activity", activity.get("activity_title")),
                sentence("Date", activity.get("activity_date")),
                sentence("Related company or fund", activity.get("related_company_or_fund")),
                sentence("Sector", activity.get("sector")),
                sentence("Geography", activity.get("geography")),
            ]
            if part
        )
    else:
        pieces.extend(
            part
            for part in [
                sentence("Activity", row.get("recent_activity_or_signal")),
                sentence("Date", row.get("recent_activity_date")),
            ]
            if part
        )
    return " ".join(pieces)


def main() -> int:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found at {WORKBOOK_PATH}")

    if not BACKUP_PATH.exists():
        shutil.copy2(WORKBOOK_PATH, BACKUP_PATH)

    workbook = load_workbook(WORKBOOK_PATH)
    family_sheet = workbook["Family_Offices"]
    normalized_source_rows = normalize_source_log_urls(workbook["Source_Log"])
    family_rows = rows_by_header(family_sheet)
    activity_rows = rows_by_header(workbook["Recent_Activities"])
    source_rows = rows_by_header(workbook["Source_Log"])

    family_by_id = {clean(row.get("record_id")): row for row in family_rows}
    activity_by_id = {clean(row.get("record_id")): row for row in activity_rows}
    sources_by_id: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in source_rows:
        sources_by_id[clean(row.get("record_id"))].append(row)

    rag_sheet = workbook["RAG_Documents"]
    headers = sheet_headers(rag_sheet)
    text_col = headers["text_for_embedding"]
    source_urls_col = headers["source_urls"]
    doc_type_col = headers["doc_type"]
    doc_id_col = headers["doc_id"]

    updated = 0
    for row_number in range(2, rag_sheet.max_row + 1):
        record_id = clean(rag_sheet.cell(row=row_number, column=headers["record_id"]).value)
        doc_type = clean(rag_sheet.cell(row=row_number, column=headers["doc_type"]).value)
        family_row = family_by_id.get(record_id)
        if not family_row:
            continue

        if doc_type == "entity_profile":
            text = build_entity_text(family_row)
        elif doc_type == "investment_profile":
            text = build_investment_text(family_row)
        elif doc_type in {"validation_profile", "source_profile"}:
            rag_sheet.cell(row=row_number, column=doc_type_col).value = "source_profile"
            doc_id = clean(rag_sheet.cell(row=row_number, column=doc_id_col).value)
            if "validation_profile" in doc_id:
                rag_sheet.cell(row=row_number, column=doc_id_col).value = doc_id.replace(
                    "validation_profile",
                    "source_profile",
                )
            text = build_source_text(family_row, sources_by_id.get(record_id, []))
        elif doc_type == "recent_activity_profile":
            text = build_activity_text(family_row, activity_by_id.get(record_id))
        else:
            continue

        rag_sheet.cell(row=row_number, column=text_col).value = text
        urls = source_urls(sources_by_id.get(record_id, []))
        if urls:
            rag_sheet.cell(row=row_number, column=source_urls_col).value = json.dumps(urls)
        updated += 1

    if "Data_Dictionary" in workbook.sheetnames:
        dictionary_sheet = workbook["Data_Dictionary"]
        dictionary_headers = sheet_headers(dictionary_sheet)
        for row_number in range(2, dictionary_sheet.max_row + 1):
            column_name = clean(dictionary_sheet.cell(row=row_number, column=dictionary_headers["column_name"]).value)
            sheet_name = clean(dictionary_sheet.cell(row=row_number, column=dictionary_headers["sheet_name"]).value)
            if column_name in {"confidence_level", "confidence_score_out_of_100"}:
                dictionary_sheet.cell(
                    row=row_number,
                    column=dictionary_headers["description"],
                ).value = "Internal data-quality field retained for operations; not shown in the production UI."
                dictionary_sheet.cell(row=row_number, column=dictionary_headers["used_for_rag"]).value = "No"
    workbook.save(WORKBOOK_PATH)
    print(f"Updated {updated} RAG document texts.")
    print(f"Normalized {normalized_source_rows} source log URLs.")
    print(f"Backup: {BACKUP_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
